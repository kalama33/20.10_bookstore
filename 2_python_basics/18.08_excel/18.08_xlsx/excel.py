from openpyxl import Workbook

name = "hello_worl.xlsx"

wb = Workbook()

sheet = wb.active

sheet["A1"] = "Hello"
aheet["B1"] = "world!"

# wb.save(filename=name)

# Creating sheets

ws1 = wb.create_sheet("Mysheet")
ws1.title = "DCI sheet"
ws2 = wb.create_sheet